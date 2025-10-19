"""
Utility functions for Excel import/export operations
"""

import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


def export_material_request_to_excel(material_request):
    """
    Export Material Request to Excel file
    
    Args:
        material_request: MaterialRequest object
        
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "طلب توريد مواد"
    
    # Styling
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
    title_font = Font(bold=True, size=14, name="Cairo")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = 'طلب توريد مواد'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header information
    ws['A3'] = 'رقم الطلب'
    ws['B3'] = material_request.request_number
    ws['D3'] = 'تاريخ تقديم الطلب'
    ws['E3'] = material_request.request_date.strftime('%Y-%m-%d') if material_request.request_date else ''
    ws['G3'] = 'اسم المشروع:'
    ws['H3'] = material_request.project.name if material_request.project else ''
    
    ws['A4'] = 'مقدم الطلب'
    ws['B4'] = material_request.requester.username if material_request.requester else ''
    ws['D4'] = 'مدير المشروع'
    ws['E4'] = material_request.project_manager.username if material_request.project_manager else ''
    
    # Table headers
    headers = [
        'م',
        'رقم البند (B.O.Q)',
        'اسم المادة والوصف',
        '',
        '',
        '',
        '',
        'الوحدة',
        '',
        'الكمية المتاحة',
        'الكمية المطلوبة',
        'تاريخ التوريد المطلوب'
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge cells for headers
    ws.merge_cells(f'C{row}:G{row}')  # اسم المادة والوصف
    ws.merge_cells(f'H{row}:I{row}')  # الوحدة
    
    # Data rows
    row = 7
    for idx, item in enumerate(material_request.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.boq_item_number or '').border = border
        
        # Merge cells for material name
        ws.merge_cells(f'C{row}:G{row}')
        cell = ws.cell(row=row, column=3, value=item.material_name)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        # Merge cells for unit
        ws.merge_cells(f'H{row}:I{row}')
        ws.cell(row=row, column=8, value=item.unit).border = border
        
        ws.cell(row=row, column=10, value=item.quantity_available).border = border
        ws.cell(row=row, column=11, value=item.quantity_requested).border = border
        ws.cell(row=row, column=12, value=item.required_date.strftime('%Y-%m-%d') if item.required_date else '').border = border
        
        row += 1
    
    # Notes section
    row += 2
    ws.cell(row=row, column=1, value='ملاحظات:')
    ws.cell(row=row+1, column=1, value=material_request.notes or '')
    
    # Signatures
    row += 4
    ws.cell(row=row, column=2, value='توقيع مقدم الطلب')
    ws.cell(row=row, column=11, value='توقيع مدير المشروع')
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_material_return_to_excel(material_return):
    """
    Export Material Return to Excel file
    
    Args:
        material_return: MaterialReturn object
        
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "مرتجع مواد"
    
    # Styling
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
    title_font = Font(bold=True, size=14, name="Cairo")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'مرتجع مواد'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header information
    ws['A3'] = 'رقم الطلب'
    ws['B3'] = material_return.return_number
    ws['D3'] = 'تاريخ تقديم الطلب'
    ws['E3'] = material_return.return_date.strftime('%Y-%m-%d') if material_return.return_date else ''
    ws['G3'] = 'اسم المشروع:'
    ws['H3'] = material_return.project.name if material_return.project else ''
    
    ws['A4'] = 'مقدم الطلب'
    ws['B4'] = material_return.requester.username if material_return.requester else ''
    ws['D4'] = 'مدير المشروع'
    ws['E4'] = material_return.project_manager.username if material_return.project_manager else ''
    
    # Table headers
    headers = [
        'م',
        'رقم البند (B.O.Q)',
        'اسم المادة والوصف',
        '',
        '',
        '',
        '',
        'الوحدة',
        '',
        'الكمية',
        'الملاحظات'
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge cells for headers
    ws.merge_cells(f'C{row}:G{row}')  # اسم المادة والوصف
    ws.merge_cells(f'H{row}:I{row}')  # الوحدة
    
    # Data rows
    row = 7
    for idx, item in enumerate(material_return.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.boq_item_number or '').border = border
        
        # Merge cells for material name
        ws.merge_cells(f'C{row}:G{row}')
        cell = ws.cell(row=row, column=3, value=item.material_name)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        # Merge cells for unit
        ws.merge_cells(f'H{row}:I{row}')
        ws.cell(row=row, column=8, value=item.unit).border = border
        
        ws.cell(row=row, column=10, value=item.quantity).border = border
        ws.cell(row=row, column=11, value=item.notes or '').border = border
        
        row += 1
    
    # Notes section
    row += 2
    ws.cell(row=row, column=1, value='ملاحظات:')
    ws.cell(row=row+1, column=1, value=material_return.notes or '')
    
    # Signatures
    row += 4
    ws.cell(row=row, column=2, value='توقيع مقدم الطلب')
    ws.cell(row=row, column=10, value='توقيع مدير المشروع')
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_material_request_from_excel(file):
    """
    Import Material Request items from Excel file
    
    Args:
        file: FileStorage object from Flask request
        
    Returns:
        list: List of dictionaries containing item data
    """
    try:
        df = pd.read_excel(file, header=5)  # Data starts at row 6 (0-indexed: 5)
        
        items = []
        for _, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('اسم المادة والوصف')):
                continue
            
            item = {
                'boq_item_number': row.get('رقم البند (B.O.Q)', ''),
                'material_name': row.get('اسم المادة والوصف', ''),
                'unit': row.get('الوحدة', ''),
                'quantity_available': float(row.get('الكمية المتاحة', 0) or 0),
                'quantity_requested': float(row.get('الكمية المطلوبة', 0) or 0),
                'required_date': pd.to_datetime(row.get('تاريخ التوريد المطلوب')).date() if pd.notna(row.get('تاريخ التوريد المطلوب')) else None
            }
            items.append(item)
        
        return items
    except Exception as e:
        raise ValueError(f"خطأ في قراءة ملف Excel: {str(e)}")


def import_material_return_from_excel(file):
    """
    Import Material Return items from Excel file
    
    Args:
        file: FileStorage object from Flask request
        
    Returns:
        list: List of dictionaries containing item data
    """
    try:
        df = pd.read_excel(file, header=5)  # Data starts at row 6 (0-indexed: 5)
        
        items = []
        for _, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('اسم المادة والوصف')):
                continue
            
            item = {
                'boq_item_number': row.get('رقم البند (B.O.Q)', ''),
                'material_name': row.get('اسم المادة والوصف', ''),
                'unit': row.get('الوحدة', ''),
                'quantity': float(row.get('الكمية', 0) or 0),
                'notes': row.get('الملاحظات', '')
            }
            items.append(item)
        
        return items
    except Exception as e:
        raise ValueError(f"خطأ في قراءة ملف Excel: {str(e)}")


def download_material_request_template():
    """
    Generate an empty Material Request template for download
    
    Returns:
        BytesIO: Excel template file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "طلب توريد مواد"
    
    # Styling
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
    title_font = Font(bold=True, size=14, name="Cairo")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = 'طلب توريد مواد'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header information
    ws['A3'] = 'رقم الطلب'
    ws['D3'] = 'تاريخ تقديم الطلب'
    ws['G3'] = 'اسم المشروع:'
    ws['A4'] = 'مقدم الطلب'
    ws['D4'] = 'مدير المشروع'
    
    # Table headers
    headers = [
        'م',
        'رقم البند (B.O.Q)',
        'اسم المادة والوصف',
        '',
        '',
        '',
        '',
        'الوحدة',
        '',
        'الكمية المتاحة',
        'الكمية المطلوبة',
        'تاريخ التوريد المطلوب'
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(f'C{row}:G{row}')
    ws.merge_cells(f'H{row}:I{row}')
    
    # Add 15 empty rows
    for i in range(1, 16):
        row = 6 + i
        ws.cell(row=row, column=1, value=i).border = border
        for col in range(2, 13):
            ws.cell(row=row, column=col).border = border
        ws.merge_cells(f'C{row}:G{row}')
        ws.merge_cells(f'H{row}:I{row}')
    
    # Notes section
    row = 23
    ws.cell(row=row, column=1, value='ملاحظات:')
    
    # Signatures
    row = 25
    ws.cell(row=row, column=2, value='توقيع مقدم الطلب')
    ws.cell(row=row, column=11, value='توقيع مدير المشروع')
    
    # Auto-fit columns
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def download_material_return_template():
    """
    Generate an empty Material Return template for download
    
    Returns:
        BytesIO: Excel template file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "مرتجع مواد"
    
    # Styling
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
    title_font = Font(bold=True, size=14, name="Cairo")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'مرتجع مواد'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header information
    ws['A3'] = 'رقم الطلب'
    ws['D3'] = 'تاريخ تقديم الطلب'
    ws['G3'] = 'اسم المشروع:'
    ws['A4'] = 'مقدم الطلب'
    ws['D4'] = 'مدير المشروع'
    
    # Table headers
    headers = [
        'م',
        'رقم البند (B.O.Q)',
        'اسم المادة والوصف',
        '',
        '',
        '',
        '',
        'الوحدة',
        '',
        'الكمية',
        'الملاحظات'
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(f'C{row}:G{row}')
    ws.merge_cells(f'H{row}:I{row}')
    
    # Add 15 empty rows
    for i in range(1, 16):
        row = 6 + i
        ws.cell(row=row, column=1, value=i).border = border
        for col in range(2, 12):
            ws.cell(row=row, column=col).border = border
        ws.merge_cells(f'C{row}:G{row}')
        ws.merge_cells(f'H{row}:I{row}')
    
    # Notes section
    row = 23
    ws.cell(row=row, column=1, value='ملاحظات:')
    
    # Signatures
    row = 25
    ws.cell(row=row, column=2, value='توقيع مقدم الطلب')
    ws.cell(row=row, column=10, value='توقيع مدير المشروع')
    
    # Auto-fit columns
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- START: الدوال المفقودة التي تسببت في خطأ BOQ ImportError ---
# يجب أن تكون هذه الدوال موجودة في ملف excel_utils.py لتتمكن ملفات الـ routes من استيرادها.

def export_boq_to_excel(items):
    """
    Generate BOQ Excel for export.
    """
    # يجب استيراد هذه الدوال داخل نطاق الدالة لتفادي مشكلة numpy/pandas
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "جدول الكميات"
    
    # Styling
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        'رقم البند', 'الوصف', 'الوحدة', 'الكمية المخططة', 'الكمية المنفذة', 
        'سعر الوحدة', 'القيمة الإجمالية', 'نسبة الإنجاز (%)', 'التصنيف'
    ]
    
    row = 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Data rows
    row = 2
    for item in items:
        # التأكد من أن الـ properties موجودة قبل الوصول إليها
        ws.cell(row=row, column=1, value=getattr(item, 'item_number', ''))
        ws.cell(row=row, column=2, value=getattr(item, 'description', ''))
        ws.cell(row=row, column=3, value=getattr(item, 'unit', ''))
        ws.cell(row=row, column=4, value=getattr(item, 'quantity', 0.0))
        ws.cell(row=row, column=5, value=getattr(item, 'executed_quantity', 0.0))
        ws.cell(row=row, column=6, value=getattr(item, 'unit_price', 0.0))
        ws.cell(row=row, column=7, value=getattr(item, 'total_price', 0.0))
        ws.cell(row=row, column=8, value=f"{getattr(item, 'completion_percentage', 0.0):.2f}%")
        ws.cell(row=row, column=9, value=getattr(item, 'category', ''))
        row += 1
    
    # Auto-fit columns
    for column in ws.columns:
        ws.column_dimensions[get_column_letter(column[0].column)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_boq_from_excel(file):
    """
    Import BOQ items from Excel file. (Simplified implementation)
    """
    import pandas as pd
    
    try:
        df = pd.read_excel(file, header=0) 
        
        items = []
        for _, row in df.iterrows():
            if pd.isna(row.get('رقم البند')):
                continue
                
            items.append({
                'item_number': str(row.get('رقم البند', '')),
                'description': str(row.get('الوصف', '')),
                'unit': str(row.get('الوحدة', '')),
                'quantity': float(row.get('الكمية المخططة', 0) or 0),
                'unit_price': float(row.get('سعر الوحدة', 0) or 0),
                'total_price': float(row.get('القيمة الإجمالية', 0) or 0),
                'category': str(row.get('التصنيف', '')),
            })
        
        return items
    except Exception as e:
        # استخدام ValueError لسهولة التعامل مع الـ traceback
        raise ValueError(f"خطأ في قراءة ملف Excel: {str(e)}")

# --- END: الدوال المفقودة ---